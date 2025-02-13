import openpyxl


def export_to_excel( data, columns, filename):
    """
    Exports the given data to an Excel .xlsx file using openpyxl.
    
    :param data: A list of dicts, e.g. [{col1: val1, col2: val2}, ...]
    :param columns: A list of column names (strings).
    :param filename: Target Excel file path (e.g. 'C:/path/output.xlsx').
    """
    # Create workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IIS Logs Export"

    # Write the header row
    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write data rows
    for row_idx, row_dict in enumerate(data, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_dict.get(col_name, ""))

    # Finally, save
    wb.save(filename)